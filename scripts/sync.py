#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import sys
import tempfile
import urllib.request
from collections import Counter, defaultdict
from dataclasses import asdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scripts.majsoul import PaipuAuthRequired, PaipuError, extract_record_id, extract_uuid, fetch_record, has_yostar_credentials, parse_record, prefetch_authenticated_records


ROOT = Path(__file__).resolve().parents[1]
NAT_CODES = {
    "chile": "CL", "chileno": "CL", "chilena": "CL",
    "uruguay": "UY", "uruguayo": "UY", "uruguaya": "UY",
    "argentina": "AR", "argentino": "AR", "argentina": "AR",
    "peru": "PE", "peruana": "PE", "peruano": "PE",
    "brasil": "BR", "brazil": "BR", "brasileiro": "BR", "brasileira": "BR",
    "mexico": "MX", "mexicano": "MX", "mexicana": "MX", "méxicana": "MX",
}
CALENDAR_VALUE_COLS = {"A": [3, 6, 9, 12, 15, 18], "B": [22, 25, 28, 31, 34, 37]}
CALENDAR_PLAYER_COLS = {"A": [1, 4, 7, 10, 13, 16], "B": [20, 23, 26, 29, 32, 35]}
SESSION_G1_ROWS = [11, 19, 27, 35, 43, 51, 59]
# Mínimo de jugadores en común para dar por equivalentes dos grupos: 3 de 4,
# para tolerar exactamente un suplente.
MIN_ROSTER_OVERLAP = 3
WEEKDAYS_ES = ["lun", "mar", "mié", "jue", "vie", "sáb", "dom"]
MONTHS_ES = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]


class SyncError(RuntimeError):
    pass


def cell_value(cell: Any) -> Any:
    value = cell.value
    if isinstance(value, str):
        return value.strip()
    return value


def json_default(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    raise TypeError(type(value).__name__)


def format_date(value: Any) -> tuple[str, str, str | None]:
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return f"{value.day:02d} {MONTHS_ES[value.month - 1]}", WEEKDAYS_ES[value.weekday()], value.isoformat()
    return "Por definir", "—", None


def normalize_nat(value: Any) -> str:
    raw = str(value or "").lower().replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    for label, code in NAT_CODES.items():
        normalized = label.replace("é", "e")
        if normalized in raw:
            return code
    return "OT"


def download_sheet(spreadsheet_id: str, destination: Path) -> None:
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    request = urllib.request.Request(url, headers={"User-Agent": "LigaMahjongChile/1.0"})
    try:
        with urllib.request.urlopen(request, timeout=45) as response, destination.open("wb") as output:
            shutil.copyfileobj(response, output)
    except Exception as exc:
        raise SyncError(f"No se pudo descargar Google Sheets: {exc}") from exc


def load_config(path: Path) -> dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise SyncError(f"Configuración inválida en {path}: {exc}") from exc


def read_roster(workbook: Any, division: str, sheet_name: str) -> list[dict[str, Any]]:
    ws = workbook[sheet_name]
    players = []
    seen_ids: set[int] = set()
    for row in range(2, ws.max_row + 1):
        number = cell_value(ws.cell(row, 1))
        name = cell_value(ws.cell(row, 2))
        account = cell_value(ws.cell(row, 3))
        if number in (None, "") or not name:
            continue
        league_id = f"{division}{int(number):02d}"
        account_id = int(account) if account not in (None, "") else None
        if account_id is not None and account_id in seen_ids:
            raise SyncError(f"{sheet_name}: ID Mahjong Soul duplicado {account_id}")
        if account_id is not None:
            seen_ids.add(account_id)
        players.append({
            "id": league_id, "div": division, "num": f"{int(number):02d}",
            "name": str(name), "shortName": str(name), "handle": str(name),
            "accountId": account_id, "discord": str(cell_value(ws.cell(row, 4)) or ""),
            "nat": normalize_nat(cell_value(ws.cell(row, 5))),
        })
    if len(players) != 24:
        raise SyncError(f"{sheet_name}: se esperaban 24 jugadores y se encontraron {len(players)}")
    return players


def read_calendar(workbook: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    ws = workbook["Calendario"]
    fixtures: list[dict[str, Any]] = []
    submissions: list[dict[str, Any]] = []
    seen_uuids: dict[str, str] = {}
    for division in ("A", "B"):
        for session, g1_row in enumerate(SESSION_G1_ROWS, start=1):
            for table_idx, (player_col, value_col) in enumerate(zip(CALENDAR_PLAYER_COLS[division], CALENDAR_VALUE_COLS[division]), start=1):
                players = [str(cell_value(ws.cell(g1_row - 2 + offset, player_col)) or "") for offset in range(4)]
                raw_date = cell_value(ws.cell(g1_row - 2, value_col))
                raw_time = cell_value(ws.cell(g1_row - 1, value_col))
                display_date, weekday, iso_date = format_date(raw_date)
                fixture = {
                    "division": division, "session": session, "table": table_idx,
                    "players": players, "date": display_date, "weekday": weekday,
                    "dateISO": iso_date, "time": raw_time.isoformat(timespec="minutes") if isinstance(raw_time, time) else None,
                }
                fixtures.append(fixture)
                for game in (1, 2):
                    cell = ws.cell(g1_row + game - 1, value_col)
                    value = str(cell_value(cell) or "")
                    key = f"{division}-S{session}-M{table_idx}-G{game}"
                    entry = {"key": key, **fixture, "game": game, "cell": f"Calendario!{get_column_letter(value_col)}{cell.row}", "url": value}
                    if value:
                        uuid = extract_uuid(value)
                        if uuid in seen_uuids:
                            raise SyncError(f"Paipu duplicado en {entry['cell']} y {seen_uuids[uuid]}: {uuid}")
                        seen_uuids[uuid] = entry["cell"]
                        entry["uuid"] = uuid
                        entry["recordId"] = extract_record_id(value)
                    submissions.append(entry)
    return fixtures, submissions


def parse_history(workbook: Any, division: str, sheet_name: str, rule: dict[str, Any]) -> dict[str, dict[str, Any]]:
    ws = workbook[sheet_name]
    output: dict[str, dict[str, Any]] = {}
    pattern = re.compile(r"^S(\d+)\s+M(\d+)\s+G(\d+)$")
    for row in range(2, 86):
        label = str(cell_value(ws.cell(row, 1)) or "")
        match = pattern.match(label)
        if not match:
            continue
        session, table, game = map(int, match.groups())
        raw = str(cell_value(ws.cell(row, 2)) or "")
        if not raw:
            continue
        parts = [part.strip() for part in raw.split(",")]
        if len(parts) != 8:
            raise SyncError(f"{sheet_name}!B{row}: el resultado debe contener 8 valores")
        results = []
        try:
            for place in range(4):
                score = int(float(parts[place * 2 + 1]))
                points = round((score - int(rule["initialPoints"])) / 1000 + float(rule["uma"][place]), 1)
                results.append({"name": parts[place * 2], "scoreRaw": score, "place": place + 1, "delta": points})
        except ValueError as exc:
            raise SyncError(f"{sheet_name}!B{row}: puntaje inválido") from exc
        if sum(item["scoreRaw"] for item in results) != 120000:
            raise SyncError(f"{sheet_name}!B{row}: los scores no suman 120.000")
        key = f"{division}-S{session}-M{table}-G{game}"
        output[key] = {"key": key, "session": session, "table": table, "game": game, "results": results, "sourceCell": f"{sheet_name}!B{row}"}
    return output


def merge_paipus(submissions: list[dict[str, Any]], histories: dict[str, dict[str, Any]], cache_dir: Path, offline: bool) -> tuple[dict[str, Any], dict[str, Any]]:
    status: list[dict[str, Any]] = []
    parsed_games: dict[str, Any] = {}
    for submission in submissions:
        if "uuid" not in submission:
            status.append({"key": submission["key"], "cell": submission["cell"], "status": "PENDIENTE", "message": "Falta paipu"})
            continue
        uuid = submission["uuid"]
        try:
            cache_file = cache_dir / f"{uuid}.pb"
            if offline and not cache_file.exists():
                raise PaipuError("Sin caché local; omite --offline para descargarlo")
            raw = cache_file.read_bytes() if cache_file.exists() else fetch_record(submission["recordId"], cache_dir, uuid)
            parsed = parse_record(uuid, raw)
            if any(not name for name in submission["players"]):
                raise PaipuError("El fixture todavía no contiene cuatro jugadores")
            # El paipu manda sobre Game History (que es un duplicado). Si el
            # registro declara la identidad de los 4 asientos, lo aceptamos
            # como fuente; si no, se conserva Game History como respaldo.
            has_players = len(parsed.players) == 4 and all(p.get("account_id") for p in parsed.players)
            state = "PUBLICADO" if has_players else "VALIDADO"
            identity_count = sum(1 for p in parsed.players if p.get("account_id"))
            message = (
                "Paipu decodificado (identidad de asientos)" if has_players
                else f"Paipu sin identidad de asientos (record_game={parsed.record_game_seen}, {identity_count}/4 con account_id)"
            )
            parsed_games[submission["key"]] = {
                "uuid": uuid, "url": submission["url"], "sha256": parsed.sha256,
                "finalScoresBySeat": parsed.final_scores, "seatStats": parsed.seat_stats,
                "players": parsed.players, "hands": parsed.hands, "status": state,
            }
            status.append({"key": submission["key"], "cell": submission["cell"], "uuid": uuid, "status": state, "message": message})
        except PaipuAuthRequired as exc:
            status.append({"key": submission["key"], "cell": submission["cell"], "uuid": uuid, "status": "REQUIERE_AUTH", "message": str(exc)})
        except Exception as exc:
            message = str(exc) if isinstance(exc, PaipuError) else f"{type(exc).__name__}: {exc}"
            status.append({"key": submission["key"], "cell": submission["cell"], "uuid": uuid, "status": "ERROR", "message": message})
    return parsed_games, {"submissions": status}


def pct(value: int, total: int) -> float:
    return round(value * 100 / total, 1) if total else 0.0


def key_sort_key(key: str) -> tuple[int, int, int]:
    match = re.match(r"^[AB]-S(\d+)-M(\d+)-G(\d+)$", key)
    return tuple(int(part) for part in match.groups()) if match else (0, 0, 0)


def match_paipu_seats(parsed_players: list[dict[str, Any]], players: list[dict[str, Any]]) -> dict[int, dict[str, Any]] | None:
    """Mapa asiento→jugador por account_id (o nickname), sin depender del orden del fixture."""
    by_account = {p["accountId"]: p for p in players if p.get("accountId") is not None}
    by_name = {p["name"].lower(): p for p in players}
    seat_map: dict[int, dict[str, Any]] = {}
    for seat, entry in enumerate(parsed_players):
        account_id = entry.get("account_id")
        player = by_account.get(account_id) if account_id is not None else None
        if player is None and entry.get("nickname"):
            player = by_name.get(str(entry["nickname"]).strip().lower())
        if player is not None:
            seat_map[seat] = player
    if len(seat_map) == len({p["id"] for p in seat_map.values()}) >= 3:
        # 4 asientos = mesa completa; 3 = hay exactamente un suplente.
        return seat_map
    return None


def match_fixture_order(fixture_players: list[str], players: list[dict[str, Any]]) -> dict[int, dict[str, Any]] | None:
    """Mapa asiento→jugador según el orden del fixture (fallback cuando el paipu no trae identidad)."""
    by_name = {p["name"].lower(): p for p in players}
    seat_map: dict[int, dict[str, Any]] = {}
    for seat, name in enumerate(fixture_players[:4]):
        player = by_name.get(name.strip().lower())
        if player is not None:
            seat_map[seat] = player
    if len(seat_map) == len({p["id"] for p in seat_map.values()}) >= 3:
        # 4 asientos = mesa completa; 3 = hay exactamente un suplente.
        return seat_map
    return None


def find_absent_player(fixture_players: list[str], players: list[dict[str, Any]], present_ids: set[str]) -> dict[str, Any] | None:
    """El jugador del grupo del calendario que no se sentó a la mesa."""
    by_name = {p["name"].strip().lower(): p for p in players}
    for name in fixture_players:
        player = by_name.get(str(name).strip().lower())
        if player is not None and player["id"] not in present_ids:
            return player
    return None


def build_paipu_results(key: str, seat_map: dict[int, dict[str, Any]], parsed_players: list[dict[str, Any]], final_scores: list[int], rule: dict[str, Any], absent: dict[str, Any] | None) -> list[dict[str, Any]]:
    """Resultados de liga (puesto, delta, puntos) a partir de los datos del paipu.
    Un asiento fuera del roster es un suplente: juega y ocupa su puesto, pero se
    marca con `sustitutoDe` para excluirlo de la clasificación."""
    pairs = []
    for seat in range(4):
        player = seat_map.get(seat)
        if player is None:
            apodo = str(parsed_players[seat].get("nickname") or "").strip() or "Suplente"
            player = {"id": None, "name": apodo, "handle": apodo, "nat": "OT"}
        pairs.append((player, final_scores[seat]))
    ranked = sorted(pairs, key=lambda item: item[1], reverse=True)
    results = []
    for place, (player, score) in enumerate(ranked, start=1):
        delta = round((score - int(rule["initialPoints"])) / 1000 + float(rule["uma"][place - 1]), 1)
        suplente = player["id"] is None
        results.append({
            "id": player["id"] or f"sub-{key}-{place}", "name": player["name"],
            "handle": player["handle"], "nat": player["nat"], "scoreRaw": int(score),
            "place": place, "delta": delta,
            "sustitutoDe": (absent["id"] if absent else None) if suplente else None,
        })
    return results


def build_excel_results(official: dict[str, Any], fixture_players: list[str], players: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Resultados de respaldo cuando la división/game no tiene paipu con identidad."""
    by_name = {p["name"].lower(): p for p in players}
    known = {item["name"].lower() for item in official["results"] if item["name"].lower() in by_name}
    missing = [name for name in fixture_players if name.lower() not in known]
    results = []
    for item in official["results"]:
        player = by_name.get(item["name"].lower())
        if player:
            results.append({
                "id": player["id"], "name": player["name"], "handle": player["handle"],
                "nat": player["nat"], "scoreRaw": item["scoreRaw"], "place": item["place"],
                "delta": item["delta"],
            })
            continue
        replaced_name = missing.pop(0) if missing else None
        replaced = by_name.get(replaced_name.lower()) if replaced_name else None
        results.append({
            "id": f"sub-{official['key']}-{item['place']}", "name": item["name"],
            "handle": item["name"], "nat": "OT", "scoreRaw": item["scoreRaw"],
            "place": item["place"], "delta": item["delta"],
            "sustitutoDe": replaced["id"] if replaced else None,
        })
    return results


def build_public_data(config: dict[str, Any], rosters: dict[str, list[dict[str, Any]]], fixtures: list[dict[str, Any]], submissions: list[dict[str, Any]], histories: dict[str, dict[str, Any]], parsed_games: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    divisions: dict[str, Any] = {}
    all_players: list[dict[str, Any]] = []
    stats_output: dict[str, Any] = {"players": {}}
    submission_by_key = {item["key"]: item for item in submissions if item.get("url")}
    absence_penalty = float(config.get("absencePenaltyPerHanchan", -30))
    for division in ("A", "B"):
        rule = config["divisions"][division]
        players = [{**player, "games": 0, "points": 0.0, "history": [], "cum": [], "counts": [0, 0, 0, 0], "absences": 0, "hands": 0, "wins": 0, "dealIns": 0, "riichis": 0, "openHands": 0, "damaten": 0, "winPoints": 0, "dealInPoints": 0, "winTurns": 0, "yakuCounts": Counter()} for player in rosters[division]]
        by_id = {player["id"]: player for player in players}
        matches = []
        keys = {key for key in histories if key.startswith(f"{division}-")} | {key for key in parsed_games if key.startswith(f"{division}-")}
        for key in sorted(keys, key=key_sort_key):
            official = histories.get(key)
            parsed = parsed_games.get(key)
            submission = submission_by_key.get(key)
            session, table, game = key_sort_key(key)
            fixture = next((item for item in fixtures if item["division"] == division and item["session"] == session and item["table"] == table), None)
            fixture_names = fixture["players"] if fixture else []
            source = "sin resultado"
            seat_map = None
            if parsed and len(parsed.get("players", [])) == 4:
                seat_map = match_paipu_seats(parsed["players"], players)
            if parsed and seat_map is None and fixture_names:
                seat_map = match_fixture_order(fixture_names, players)
            if parsed and seat_map and len(parsed["finalScoresBySeat"]) == 4:
                absent = None
                if len(seat_map) < 4:
                    absent = find_absent_player(fixture_names, players, {p["id"] for p in seat_map.values()})
                results = build_paipu_results(key, seat_map, parsed["players"], parsed["finalScoresBySeat"], rule, absent)
                source = "paipu"
            elif official:
                results = build_excel_results(official, fixture_names, players)
                source = "excel"
            else:
                continue
            match_players = []
            for result in results:
                es_suplente = result["id"].startswith("sub-")
                player = by_id.get(result["id"])
                if player and not es_suplente:
                    player["games"] += 1
                    player["points"] = round(player["points"] + result["delta"], 1)
                    player["history"].append(result["delta"])
                    player["counts"][result["place"] - 1] += 1
                if es_suplente and result.get("sustitutoDe"):
                    # El ausente tiene la partida por jugada y recibe la
                    # penalización fija; no ocupa puesto porque no jugó.
                    ausente = by_id.get(result["sustitutoDe"])
                    if ausente is not None:
                        ausente["games"] += 1
                        ausente["absences"] += 1
                        ausente["points"] = round(ausente["points"] + absence_penalty, 1)
                        ausente["history"].append(absence_penalty)
                match_players.append({k: result[k] for k in ("id", "name", "handle", "nat", "scoreRaw", "place", "delta")} | {"sustitutoDe": result.get("sustitutoDe"), "esSuplente": es_suplente})
            if parsed and seat_map:
                for seat, player in seat_map.items():
                    seat_stats = parsed["seatStats"][seat]
                    for field in ("hands", "wins", "dealIns", "riichis", "openHands", "damaten", "winPoints", "dealInPoints", "winTurns"):
                        player[field] += int(seat_stats[field])
                    player["yakuCounts"].update(seat_stats["yaku"])
            date_display = fixture["date"] if fixture else "—"
            matches.append({
                "id": key, "code": key, "div": division, "session": session,
                "sessionCode": f"S{session}", "hanchan": game,
                "date": date_display, "weekday": fixture["weekday"] if fixture else "—",
                "table": table, "players": match_players,
                "paipuUrl": submission["url"] if submission else None,
                "verified": source == "paipu", "source": source,
            })
        for player in players:
            running = 0.0
            player["cum"] = []
            for delta in player["history"]:
                running = round(running + delta, 1)
                player["cum"].append(running)
            games = player["games"]
            # Las ausencias suman partidas y puntos, pero no tienen puesto: el
            # promedio y la distribución se calculan sobre lo realmente jugado.
            played = games - player["absences"]
            player["placements"] = {f"p{i + 1}": round(player["counts"][i] / played, 2) if played else 0 for i in range(4)}
            player["avgRank"] = round(sum((i + 1) * count for i, count in enumerate(player["counts"])) / played, 2) if played else 0
            player["avgPoints"] = round(player["points"] / games, 1) if games else 0
            player["streak"] = round(sum(player["history"][-4:]), 1)
            hands = player["hands"]
            player["winRate"] = pct(player["wins"], hands)
            player["dealInRate"] = pct(player["dealIns"], hands)
            player["riichiRate"] = pct(player["riichis"], hands)
            player["openRate"] = pct(player["openHands"], hands)
            # Métricas al estilo amae-koromo: las tres primeras se miden sobre
            # manos ganadas o sobre deal-ins, no sobre el total de manos.
            wins, deal_ins = player["wins"], player["dealIns"]
            player["damatenRate"] = pct(player["damaten"], wins)
            player["avgWinPoints"] = round(player["winPoints"] / wins) if wins else 0
            player["avgDealInPoints"] = round(player["dealInPoints"] / deal_ins) if deal_ins else 0
            player["avgWinTurn"] = round(player["winTurns"] / wins, 2) if wins else 0
            player["topYaku"] = [{"name": name, "count": count} for name, count in player["yakuCounts"].most_common(5)]
            player["statsSample"] = hands
            player["statsReliable"] = hands >= int(config["minimumAdvancedStatsHands"])
            player["arch"] = "con datos" if player["statsReliable"] else "stats pendientes"
            stats_output["players"][player["id"]] = {key: player[key] for key in ("hands", "wins", "dealIns", "winRate", "dealInRate", "riichiRate", "openRate",
                "damatenRate", "avgWinPoints", "avgDealInPoints", "avgWinTurn", "topYaku", "statsReliable")}
            del player["yakuCounts"]
        players.sort(key=lambda item: (-item["points"], item["avgRank"] if item["games"] else 99, item["name"].lower()))
        for index, player in enumerate(players, start=1):
            player["rank"] = index
            player["zone"] = ("title" if index <= 4 else "relegation" if index >= 21 else None) if division == "A" else ("promotion" if index <= 4 else "bottom" if index >= 21 else None)
        session_items = []
        for session in range(1, int(config["sessionsTotal"]) + 1):
            session_matches = [match for match in matches if match["session"] == session]
            fixture = next((item for item in fixtures if item["division"] == division and item["session"] == session and item["dateISO"]), None)
            session_items.append({"n": session, "code": f"S{session}", "date": fixture["date"] if fixture else "Por definir", "weekday": fixture["weekday"] if fixture else "—", "div": division, "matches": len(session_matches), "status": "played" if len(session_matches) == 12 else "partial" if session_matches else "pending"})
        divisions[division] = {"key": division, "players": players, "matches": matches, "sessions": session_items}
        all_players.extend(players)

    sessions_played = min(sum(1 for s in divisions[d]["sessions"] if s["status"] == "played") for d in ("A", "B"))
    next_session_number = min(sessions_played + 1, int(config["sessionsTotal"]))
    next_fixture = next((f for f in fixtures if f["session"] == next_session_number and f["dateISO"]), None)
    next_session = {"code": f"S{next_session_number}", "date": next_fixture["date"] if next_fixture else "Por definir", "day": next_fixture["weekday"] if next_fixture else "—"}
    chile_a = [p for p in divisions["A"]["players"] if p["nat"] == "CL"]
    for index, player in enumerate(chile_a, start=1):
        player["natRank"] = index
        player["iormc"] = "qualified" if index <= 4 else "contention" if index <= 7 else "out"
    cut = chile_a[3] if len(chile_a) > 3 else None
    bubble = chile_a[4] if len(chile_a) > 4 else None
    iormc = {"slots": 4, "eligible": len(chile_a), "qualified": chile_a[:4], "contention": chile_a[4:7], "rest": chile_a[7:], "all": chile_a, "cutPoints": cut["points"] if cut else 0, "gap": round(cut["points"] - bubble["points"], 1) if cut and bubble else 0}
    nationalities = []
    for code in sorted({p["nat"] for p in all_players}):
        group = [p for p in all_players if p["nat"] == code]
        best = max(group, key=lambda p: p["points"])
        nationalities.append({"code": code, "count": len(group), "inA": sum(p["div"] == "A" for p in group), "inB": sum(p["div"] == "B" for p in group), "avgPoints": round(sum(p["points"] for p in group) / len(group), 1), "avgRank": round(sum(p["avgRank"] for p in group) / len(group), 2), "best": best})
    player_nat = {p["name"].lower(): p["nat"] for p in all_players}
    calendar = []
    for fixture in fixtures:
        if fixture["session"] < next_session_number:
            continue
        calendar.append({
            "date": fixture["date"], "day": fixture["weekday"],
            "round": f"Sesión {fixture['session']}", "session": fixture["session"],
            "table": fixture["table"],
            "mesa": f"División {fixture['division']} — Mesa {fixture['table']}",
            "time": fixture["time"] or "Por definir",
            "div": fixture["division"],
            "players": [{"name": n, "nat": player_nat.get(n.lower(), "OT")} for n in fixture["players"]],
            "status": "highlight" if fixture["session"] == next_session_number else "scheduled",
        })
    data = {
        "divisions": divisions, "allPlayers": all_players, "nationalities": nationalities,
        "iormc": iormc, "calendar": calendar,
        "league": {"season": config["seasonLabel"], "sessionsPlayed": sessions_played, "sessionsTotal": int(config["sessionsTotal"]), "hanchanPerSession": 2, "playersPerDiv": 24, "hanchanPerDiv": max(len(divisions["A"]["matches"]), len(divisions["B"]["matches"])), "hanchanTotal": len(divisions["A"]["matches"]) + len(divisions["B"]["matches"]), "nextSession": next_session, "rules": {key: {"initialPoints": value["initialPoints"], "uma": value["uma"]} for key, value in config["divisions"].items()}},
    }
    add_hall_of_fame(data)
    return data, stats_output


def add_hall_of_fame(data: dict[str, Any]) -> None:
    for division in ("A", "B"):
        players = data["divisions"][division]["players"]
        top = players[0]
        def best(field: str, lower: bool = False) -> dict[str, Any]:
            candidates = [p for p in players if p["games"] > 0]
            return (min if lower else max)(candidates or players, key=lambda p: p[field])
        records = [
            {"tag": "Líder División", "value": f"{top['points']:+.1f}", "sub": "puntos uma", "player": top, "jp": "王座"},
            {"tag": "Mejor Win Rate", "value": f"{best('winRate')['winRate']:.1f}%", "sub": "manos ganadas", "player": best("winRate"), "jp": "和了率"},
            {"tag": "Muro de Hierro", "value": f"{best('dealInRate', True)['dealInRate']:.1f}%", "sub": "deal-in más bajo", "player": best("dealInRate", True), "jp": "放銃"},
            {"tag": "Velocidad", "value": f"{best('riichiRate')['riichiRate']:.1f}%", "sub": "riichi rate", "player": best("riichiRate"), "jp": "立直"},
            {"tag": "Consistencia", "value": f"{best('avgRank', True)['avgRank']:.2f}", "sub": "puesto promedio", "player": best("avgRank", True), "jp": "平均順位"},
            {"tag": "Racha Caliente", "value": f"{best('streak')['streak']:+.1f}", "sub": "últimas 4 hanchan", "player": best("streak"), "jp": "連勝"},
        ]
        data["divisions"][division]["hallOfFame"] = records


def advanced_stats_health(stats: dict[str, Any], status: dict[str, Any]) -> dict[str, Any]:
    submissions = status.get("submissions", [])
    counter = Counter(item.get("status") for item in submissions)
    players_map = stats["players"] if "players" in stats else stats
    players = list(players_map.values())
    reliable = sum(1 for p in players if p.get("statsReliable"))
    with_hands = sum(1 for p in players if (p.get("hands") or 0) > 0)
    issues = [
        {"status": item.get("status"), "cell": item.get("cell"), "message": item.get("message", "")}
        for item in submissions if item.get("status") in ("REQUIERE_AUTH", "ERROR")
    ]
    published = counter.get("PUBLICADO", 0)
    validated = counter.get("VALIDADO", 0)
    pendiente = counter.get("PENDIENTE", 0)
    requiere_auth = counter.get("REQUIERE_AUTH", 0)
    errores = counter.get("ERROR", 0)
    submitted = len(submissions) - pendiente
    summary = (
        f"{with_hands}/{len(players)} jugadores con datos avanzados; "
        f"paipus: {published} publicado, {validated} validado, "
        f"{requiere_auth} requiere auth, {errores} error, "
        f"{pendiente} pendiente"
    )
    return {
        "reliable": reliable, "with_hands": with_hands, "issues": issues,
        "publicado": published, "validado": validated,
        "requiere_auth": requiere_auth, "errores": errores,
        "pendiente": pendiente, "submitted": submitted,
        "summary": summary,
    }


def align_history_with_fixtures(histories: dict[str, dict[str, Any]], fixtures: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], list[str]]:
    """El Game History numera las mesas distinto que el Calendario: la sesión
    coincide, el número de mesa no. Se empareja cada grupo del historial con la
    mesa del calendario con la que comparta al menos 3 jugadores (3 y no 4 para
    tolerar un suplente) y se re-indexa a la mesa real. El calendario manda."""
    avisos: list[str] = []
    nombres_gh: dict[tuple[str, int, int], set[str]] = {}
    claves_gh: dict[tuple[str, int, int], list[str]] = defaultdict(list)
    for key, official in histories.items():
        division = key.split("-", 1)[0]
        celda = (division, official["session"], official["table"])
        nombres_gh.setdefault(celda, set()).update(
            r["name"].strip().lower() for r in official["results"]
        )
        claves_gh[celda].append(key)

    remapeo: dict[tuple[str, int, int], int] = {}
    sesiones = {(d, s) for d, s, _ in nombres_gh}
    for division, session in sorted(sesiones):
        del_calendario = {
            f["table"]: {n.strip().lower() for n in f["players"] if n}
            for f in fixtures if f["division"] == division and f["session"] == session
        }
        mesas_gh = sorted(m for d, s, m in nombres_gh if d == division and s == session)
        # Todas las coincidencias posibles, de mayor a menor, asignadas sin repetir.
        candidatos = sorted(
            (
                (len(nombres_gh[(division, session, mesa)] & grupo), mesa, mesa_real)
                for mesa in mesas_gh
                for mesa_real, grupo in del_calendario.items()
                if len(nombres_gh[(division, session, mesa)] & grupo) >= MIN_ROSTER_OVERLAP
            ),
            reverse=True,
        )
        usadas_gh: set[int] = set()
        usadas_reales: set[int] = set()
        for coincidencias, mesa, mesa_real in candidatos:
            if mesa in usadas_gh or mesa_real in usadas_reales:
                continue
            usadas_gh.add(mesa)
            usadas_reales.add(mesa_real)
            remapeo[(division, session, mesa)] = mesa_real
        for mesa in mesas_gh:
            if mesa not in usadas_gh:
                avisos.append(
                    f"División {division}, sesión {session}, mesa {mesa} del Game History: "
                    f"ningún grupo del calendario coincide en {MIN_ROSTER_OVERLAP} jugadores; "
                    f"se conserva el número de mesa original"
                )
                remapeo[(division, session, mesa)] = mesa

    realineadas: dict[str, dict[str, Any]] = {}
    for celda, keys in claves_gh.items():
        division, session, mesa = celda
        mesa_real = remapeo.get(celda, mesa)
        for key in keys:
            official = dict(histories[key])
            official["table"] = mesa_real
            official["tableGameHistory"] = mesa
            nueva = f"{division}-S{session}-M{mesa_real}-G{official['game']}"
            official["key"] = nueva
            if nueva in realineadas:
                avisos.append(
                    f"División {division}, sesión {session}: dos grupos del Game History "
                    f"apuntan a la mesa {mesa_real}; se descarta el duplicado {key}"
                )
                continue
            realineadas[nueva] = official
            if mesa_real != mesa:
                avisos.append(
                    f"División {division}, sesión {session}: el grupo de la mesa {mesa} del "
                    f"Game History corresponde a la mesa {mesa_real} del calendario"
                )
    return realineadas, avisos


def write_outputs(data: dict[str, Any], stats: dict[str, Any], status: dict[str, Any], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for name, payload in (("liga.json", data), ("stats.json", stats), ("sync-status.json", status)):
        (output_dir / name).write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=json_default) + "\n", encoding="utf-8")
    js = "// Generado por scripts/sync.py; no editar manualmente.\nwindow.MJC_DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":"), default=json_default) + ";\n"
    (output_dir / "generated.js").write_text(js, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Sincroniza Liga Mahjong Chile desde Google Sheets y paipus")
    parser.add_argument("--config", type=Path, default=ROOT / "sync-config.json")
    parser.add_argument("--xlsx", type=Path, help="Usa una copia local del Sheet en vez de descargarlo")
    parser.add_argument("--offline", action="store_true", help="No descarga paipus; usa solo data/raw-paipu")
    parser.add_argument("--strict-paipu", action="store_true", help="Falla si algún paipu presente no se puede validar")
    parser.add_argument("--require-stats", action="store_true", help="Falla si ningún jugador tiene estadísticas avanzadas confiables")
    parser.add_argument("--output", type=Path, default=ROOT / "data")
    args = parser.parse_args()
    config = load_config(args.config)
    temp_path: Path | None = None
    try:
        if args.xlsx:
            source = args.xlsx
        else:
            handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            handle.close()
            temp_path = Path(handle.name)
            download_sheet(config["spreadsheetId"], temp_path)
            source = temp_path
        workbook = load_workbook(source, data_only=False)
        rosters = {division: read_roster(workbook, division, rule["rosterSheet"]) for division, rule in config["divisions"].items()}
        fixtures, submissions = read_calendar(workbook)
        histories: dict[str, dict[str, Any]] = {}
        for division, rule in config["divisions"].items():
            histories.update(parse_history(workbook, division, rule["historySheet"], rule))
        histories, avisos_mesas = align_history_with_fixtures(histories, fixtures)
        for aviso in avisos_mesas:
            print(f"AVISO: {aviso}", file=sys.stderr)
        if not args.offline:
            # Una falla de la sesión técnica no debe frenar la publicación de
            # standings; los paipus quedan REQUIERE_AUTH y se reintenta luego.
            submitted = sum(1 for item in submissions if item.get("uuid"))
            if submitted and not has_yostar_credentials():
                print("AVISO: hay paipus que requieren sesión técnica pero no hay credenciales MAJSOUL_UID/TOKEN/DEVICE_ID. Las estadísticas avanzadas quedarán como pendientes.", file=sys.stderr)
            try:
                downloaded = prefetch_authenticated_records(submissions, ROOT / "data" / "raw-paipu")
                if downloaded:
                    print(f"Paipus descargados con la sesión técnica: {downloaded}")
            except PaipuError as exc:
                print(f"AVISO: la descarga autenticada de paipus falló: {exc}", file=sys.stderr)
        parsed_games, status = merge_paipus(submissions, histories, ROOT / "data" / "raw-paipu", args.offline)
        if args.strict_paipu:
            failures = [item for item in status["submissions"] if item["status"] == "ERROR"]
            if failures:
                raise SyncError("Paipus con error:\n" + "\n".join(f"- {item['cell']}: {item['message']}" for item in failures))
        data, stats = build_public_data(config, rosters, fixtures, submissions, histories, parsed_games)
        health = advanced_stats_health(stats, status)
        if args.require_stats and health["submitted"] and not health["with_hands"]:
            raise SyncError("Estadísticas avanzadas ausentes: hay paipus enviados pero ninguno aportó manos (REQUIERE_AUTH/ERROR). Revisa la sesión técnica.")
        if (health["submitted"] and not health["with_hands"]) or health["requiere_auth"] or health["errores"]:
            print(f"AVISO: estadísticas avanzadas incompletas — {health['summary']}", file=sys.stderr)
            for issue in health["issues"]:
                print(f"  {issue['status']} en {issue['cell']}: {issue['message']}", file=sys.stderr)
        write_outputs(data, stats, status, args.output)
        counts = Counter(item["status"] for item in status["submissions"])
        print(f"Sincronización lista: {len(histories)} resultados oficiales, {len(parsed_games)} paipus procesados")
        print("Estados paipu: " + ", ".join(f"{key}={value}" for key, value in sorted(counts.items())))
        print(f"Stats avanzadas: {health['summary']}")
        print(f"Salida: {args.output}")
        return 0
    except (SyncError, PaipuError, KeyError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    finally:
        if temp_path:
            temp_path.unlink(missing_ok=True)


if __name__ == "__main__":
    raise SystemExit(main())
